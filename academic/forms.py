from django import forms
from django.core.exceptions import ValidationError
import openpyxl
from accounts.models import User
from academic.models import Class, StudentEnrollment
import secrets
import string


def generate_teacher_password(first_name, last_name, sequence_number):
    """
    Generate password dari nama guru
    Format: FirstName@guru + sequence_number
    Contoh: Ahmad@guru1
    """
    first_name_formatted = first_name.strip().capitalize()
    password = f"{first_name_formatted}@guru{sequence_number}"
    return password


def generate_student_password(first_name, middle_name, sequence_number):
    """
    Generate password dari nama siswa
    Format: MiddleName@siswa + sequence_number
    Contoh: AZZAM@siswa1
    """
    if not middle_name:
        middle_name = first_name

    # Capitalize only first letter, rest lowercase
    middle_name_formatted = middle_name.strip().capitalize()
    password = f"{middle_name_formatted}@siswa{sequence_number}"
    return password


class BulkTeacherEnrollmentForm(forms.Form):
    """Form untuk bulk enrollment guru dari Excel"""
    excel_file = forms.FileField(
        label='File Excel (.xlsx)',
        help_text='Format: Email, Nama Depan, Nama Belakang, NIP (ID Guru)',
        widget=forms.FileInput(attrs={
            'class': 'w-full px-4 py-2 border border-gray-300 rounded-lg',
            'accept': '.xlsx,.xls'
        })
    )

    def clean_excel_file(self):
        file = self.cleaned_data['excel_file']
        if not file.name.endswith(('.xlsx', '.xls')):
            raise ValidationError('File harus berformat Excel (.xlsx atau .xls)')
        return file

    def process_excel(self):
        """Process Excel file and create teachers"""
        excel_file = self.cleaned_data['excel_file']

        # Read Excel file
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        created_accounts = []
        failed_rows = []
        sequence_number = 1

        # Skip header row (row 1)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                email, first_name, last_name, nip = row[0], row[1], row[2], row[3]

                if not all([email, first_name, last_name]):
                    failed_rows.append((row_idx, 'Missing required fields'))
                    continue

                # Generate password using the format
                password = generate_teacher_password(first_name, last_name, sequence_number)

                # Create user with TEACHER role
                username = email.split('@')[0]
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'username': username,
                        'first_name': first_name,
                        'last_name': last_name,
                        'role': 'TEACHER',  # Set role as TEACHER
                    }
                )

                if created:
                    user.set_password(password)
                    user.save()

                    # Create or update UserProfile with NIP
                    from accounts.models import UserProfile
                    profile, _ = UserProfile.objects.get_or_create(user=user)
                    if nip:
                        profile.nip = nip
                        profile.save()

                created_accounts.append({
                    'email': email,
                    'name': f'{first_name} {last_name}',
                    'username': username,
                    'password': password if created else 'Existing account',
                    'nip': nip or 'N/A'
                })

                sequence_number += 1

            except Exception as e:
                failed_rows.append((row_idx, str(e)))

        return {
            'success': len(created_accounts),
            'failed': len(failed_rows),
            'accounts': created_accounts,
            'errors': failed_rows
        }


class BulkStudentEnrollmentForm(forms.Form):
    """Form untuk bulk enrollment siswa dari Excel"""
    excel_file = forms.FileField(
        label='File Excel (.xlsx)',
        help_text='Format: Email, First Name, Last Name, NIS (student ID)',
        widget=forms.FileInput(attrs={
            'class': 'w-full px-4 py-2 border border-gray-300 rounded-lg',
            'accept': '.xlsx,.xls'
        })
    )
    class_obj = forms.ModelChoiceField(
        queryset=Class.objects.all(),
        label='Kelas Tujuan',
        widget=forms.Select(attrs={
            'class': 'w-full px-4 py-2 border border-gray-300 rounded-lg'
        })
    )

    def clean_excel_file(self):
        file = self.cleaned_data['excel_file']
        if not file.name.endswith(('.xlsx', '.xls')):
            raise ValidationError('File harus berformat Excel (.xlsx atau .xls)')
        return file

    def process_excel(self):
        """Process Excel file and create students + enrollments"""
        excel_file = self.cleaned_data['excel_file']
        class_obj = self.cleaned_data['class_obj']

        # Read Excel file
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active

        created_accounts = []
        failed_rows = []
        sequence_number = 1

        # Skip header row (row 1)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                email, first_name, last_name, nis = row[0], row[1], row[2], row[3]

                if not all([email, first_name, last_name]):
                    failed_rows.append((row_idx, 'Missing required fields'))
                    continue

                # Extract middle name from full name
                # Full name format: FIRST_NAME MIDDLE_NAME LAST_NAME
                # We need to extract the middle name
                full_name_parts = [first_name, last_name]
                
                # Split first_name to check for middle name pattern
                first_name_parts = str(first_name).strip().split()
                if len(first_name_parts) > 1:
                    # If first_name contains multiple words, treat the last part as actual first name
                    # and others as middle names
                    middle_name = ' '.join(first_name_parts[:-1])
                    first_only = first_name_parts[-1]
                    # For password, use the second part (middle name)
                    password_name = first_name_parts[1] if len(first_name_parts) > 1 else first_only
                else:
                    # If only one part, use it as first name
                    first_only = first_name
                    middle_name = last_name if last_name else first_name
                    password_name = first_name

                # Generate password using the new format
                password = generate_student_password(password_name, password_name, sequence_number)

                # Create user
                username = email.split('@')[0]
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'username': username,
                        'first_name': first_name,
                        'last_name': last_name,
                        'role': 'STUDENT',
                    }
                )

                if created:
                    user.set_password(password)
                    user.save()

                # Create enrollment
                enrollment, _ = StudentEnrollment.objects.get_or_create(
                    student=user,
                    class_obj=class_obj,
                    defaults={'status': 'ACTIVE', 'student_id_in_class': nis or ''}
                )

                created_accounts.append({
                    'email': email,
                    'name': f'{first_name} {last_name}',
                    'username': username,
                    'password': password if created else 'Existing account',
                    'nis': nis or 'N/A'
                })
                
                sequence_number += 1

            except Exception as e:
                failed_rows.append((row_idx, str(e)))

        return {
            'success': len(created_accounts),
            'failed': len(failed_rows),
            'accounts': created_accounts,
            'errors': failed_rows
        }


class StudentCredentialExportForm(forms.Form):
    """Form untuk export credentials siswa ke PDF"""
    class_obj = forms.ModelChoiceField(
        queryset=Class.objects.all(),
        label='Pilih Kelas',
        widget=forms.Select(attrs={
            'class': 'w-full px-4 py-2 border border-gray-300 rounded-lg'
        })
    )
    format_choice = forms.ChoiceField(
        choices=[('pdf', 'PDF'), ('csv', 'CSV Excel')],
        label='Format Export',
        widget=forms.RadioSelect(attrs={
            'class': 'h-4 w-4'
        })
    )
