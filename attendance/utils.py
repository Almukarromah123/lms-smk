"""Attendance utility functions for exporting to PDF and XLSX formats"""

import io
import base64
import qrcode
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from django.conf import settings
from PIL import Image


# ============ QR CODE GENERATION ============

def generate_qr_code_image(data, size=250):
    """
    Generate QR code image from data string.

    Args:
        data (str): Data to encode in QR code
        size (int): Size of QR code in pixels

    Returns:
        PIL.Image: QR code image
    """
    try:
        # Create QR code using qrcode library
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=2,
        )
        qr.add_data(data)
        qr.make(fit=True)

        # Create image
        qr_image = qr.make_image(fill_color="black", back_color="white")

        # Resize to desired size
        qr_image = qr_image.resize((size, size), Image.Resampling.LANCZOS)

        return qr_image
    except Exception as e:
        raise Exception(f"Failed to generate QR code: {str(e)}")


def get_qr_code_base64(data, size=250):
    """
    Generate QR code and return as base64 encoded PNG data URI.

    Args:
        data (str): Data to encode in QR code
        size (int): Size of QR code in pixels

    Returns:
        str: Data URI with base64 encoded PNG
    """
    qr_image = generate_qr_code_image(data, size)

    # Convert PIL Image to base64
    img_buffer = io.BytesIO()
    qr_image.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    img_base64 = base64.b64encode(img_buffer.getvalue()).decode()

    return f"data:image/png;base64,{img_base64}"


# ============ DAILY STATS EXPORT ============

def export_attendance_to_pdf_daily(class_obj, date, records, subject=None):
    """Generate fancy PDF for daily attendance stats with professional styling (per-subject if provided)"""
    response = HttpResponse(content_type='application/pdf')
    subject_str = f"_{subject.name.replace(' ', '_')}" if subject else ""
    filename = f"attendance_daily_{class_obj.name.replace(' ', '_')}{subject_str}_{date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.whitesmoke,
        alignment=TA_CENTER,
        spaceAfter=12,
    )

    info_style = ParagraphStyle(
        'CustomInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        leading=14,
    )

    # Header section
    header_data = [['DAILY ATTENDANCE REPORT']]
    header_table = Table(header_data, colWidths=[7.5*inch])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('PADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.15*inch))

    # Info section
    elements.append(Paragraph(f"<b>School:</b> SMK IT AL - MUKARROMAH", info_style))
    elements.append(Paragraph(f"<b>Class:</b> {class_obj.name}", info_style))
    if subject:
        elements.append(Paragraph(f"<b>Subject:</b> {subject.name}", info_style))
    elements.append(Paragraph(f"<b>Date:</b> {date.strftime('%A, %d %B %Y')}", info_style))
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d %B %Y at %H:%M:%S')}", info_style))
    elements.append(Spacer(1, 0.15*inch))

    # Summary cards
    total = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()
    permission_count = records.filter(status='PERMISSION').count()
    sick_count = records.filter(status='SICK').count()
    attendance_pct = (present_count / total * 100) if total > 0 else 0

    summary_data = [
        ['Total Students', 'Present', 'Absent', 'Permission', 'Sick', 'Attendance %'],
        [str(total), str(present_count), str(absent_count), str(permission_count), str(sick_count), f"{attendance_pct:.1f}%"],
    ]
    summary_table = Table(summary_data, colWidths=[1.5*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.2*inch))

    # Detail table
    table_data = [['No.', 'Student Name', 'Status', 'Arrival Time']]
    status_color_map = {
        'PRESENT': colors.HexColor('#10b981'),
        'ABSENT': colors.HexColor('#ef4444'),
        'PERMISSION': colors.HexColor('#6366f1'),
        'SICK': colors.HexColor('#f59e0b'),
    }

    for i, rec in enumerate(records.order_by('student__first_name', 'student__last_name'), 1):
        table_data.append([
            str(i),
            rec.student.get_full_name(),
            rec.get_status_display(),
            rec.arrival_time.strftime('%H:%M') if rec.arrival_time else '-'
        ])

    detail_table = Table(table_data, colWidths=[0.7*inch, 3.2*inch, 1.2*inch, 1.4*inch])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 0.2*inch))

    # Footer with signature line
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER,
    )
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Teacher: __________________ | Date: __________________", footer_style))

    doc.build(elements)
    return response


def export_attendance_to_xlsx_daily(class_obj, date, records, subject=None):
    """Generate fancy XLSX for daily attendance stats with professional styling"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_daily_{class_obj.name.replace(' ', '_')}_{date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = Workbook()
    ws = workbook.active
    ws.title = f'Daily {date.strftime("%Y-%m-%d")}'

    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'DAILY ATTENDANCE REPORT'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Info section
    ws['A3'] = 'School:'
    ws['B3'] = 'SMK IT AL - MUKARROMAH'
    ws['A3'].font = Font(bold=True, size=10)

    ws['A4'] = 'Class:'
    ws['B4'] = class_obj.name
    ws['A4'].font = Font(bold=True, size=10)

    ws['A5'] = 'Date:'
    ws['B5'] = date.strftime('%A, %d %B %Y')
    ws['A5'].font = Font(bold=True, size=10)

    # Summary section
    summary_fill = PatternFill(start_color='4f46e5', end_color='4f46e5', fill_type='solid')
    summary_font = Font(bold=True, color='FFFFFF', size=10)
    summary_alignment = Alignment(horizontal='center', vertical='center')

    ws['A7'] = 'Total Students'
    ws['B7'] = 'Present'
    ws['C7'] = 'Absent'
    ws['D7'] = 'Permission'
    ws['E7'] = 'Sick'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}7'].fill = summary_fill
        ws[f'{col}7'].font = summary_font
        ws[f'{col}7'].alignment = summary_alignment

    total = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()
    permission_count = records.filter(status='PERMISSION').count()
    sick_count = records.filter(status='SICK').count()

    ws['A8'] = total
    ws['B8'] = present_count
    ws['C8'] = absent_count
    ws['D8'] = permission_count
    ws['E8'] = sick_count

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}8'].alignment = Alignment(horizontal='center')

    # Table headers
    table_header_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
    table_header_font = Font(bold=True, color='FFFFFF', size=10)
    table_header_alignment = Alignment(horizontal='center', vertical='center')

    headers = ['No.', 'Student Name', 'Status', 'Arrival Time']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = table_header_alignment

    # Data rows with styling
    status_colors = {
        'PRESENT': 'dcfce7',  # Light green
        'ABSENT': 'fee2e2',   # Light red
        'PERMISSION': 'e0e7ff',
        'SICK': 'fef3c7',     # Light amber
    }

    border_thin = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb'),
    )

    for row_idx, rec in enumerate(records.order_by('student__first_name', 'student__last_name'), 11):
        ws.cell(row=row_idx, column=1, value=row_idx - 10)
        ws.cell(row=row_idx, column=2, value=rec.student.get_full_name())
        ws.cell(row=row_idx, column=3, value=rec.get_status_display())
        ws.cell(row=row_idx, column=4, value=rec.arrival_time.strftime('%H:%M') if rec.arrival_time else '-')

        status = rec.status
        color = status_colors.get(status, 'ffffff')

        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.border = border_thin
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal='center')

    workbook.save(response)
    return response


# ============ SEMESTER REPORT EXPORT ============

def export_attendance_to_pdf_semester(class_obj, semester, summaries, start_date, end_date, subject=None):
    """Generate fancy PDF for semester attendance report"""
    response = HttpResponse(content_type='application/pdf')
    filename = f"attendance_semester_{class_obj.name.replace(' ', '_')}_sem{semester}_{end_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    info_style = ParagraphStyle(
        'CustomInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=6,
        leading=14,
    )

    # Header section
    header_data = [['SEMESTER ATTENDANCE REPORT']]
    header_table = Table(header_data, colWidths=[7.5*inch])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('PADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.15*inch))

    # Info section
    semester_label = f"Semester {semester} ({start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')})"
    elements.append(Paragraph(f"<b>School:</b> SMK IT AL - MUKARROMAH", info_style))
    elements.append(Paragraph(f"<b>Class:</b> {class_obj.name}", info_style))
    elements.append(Paragraph(f"<b>Period:</b> {semester_label}", info_style))
    elements.append(Paragraph(f"<b>Total Students:</b> {len(summaries)}", info_style))
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d %B %Y at %H:%M:%S')}", info_style))
    elements.append(Spacer(1, 0.15*inch))

    # Detail table
    table_data = [['No.', 'Student Name', 'Total Days', 'Present', 'Absent', 'Sick', 'Attendance %']]

    for i, summary in enumerate(summaries, 1):
        table_data.append([
            str(i),
            summary['student'].get_full_name(),
            str(summary['total_days']),
            str(summary['present']),
            str(summary['absent']),
            str(summary['sick']),
            f"{summary['percentage']:.1f}%"
        ])

    detail_table = Table(table_data, colWidths=[0.6*inch, 2.8*inch, 1*inch, 0.9*inch, 0.9*inch, 0.8*inch, 1.2*inch])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 0.3*inch))

    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER,
    )
    elements.append(Paragraph(f"Teacher: __________________ | Headmaster: __________________", footer_style))

    doc.build(elements)
    return response


def export_attendance_to_xlsx_semester(class_obj, semester, summaries, start_date, end_date, subject=None):
    """Generate fancy XLSX for semester attendance report"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_semester_{class_obj.name.replace(' ', '_')}_sem{semester}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = Workbook()
    ws = workbook.active
    ws.title = f'Semester {semester}'

    # Set column widths
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 12

    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'SEMESTER ATTENDANCE REPORT'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Info section
    ws['A3'] = 'School:'
    ws['B3'] = 'SMK IT AL - MUKARROMAH'
    ws['A3'].font = Font(bold=True, size=10)

    ws['A4'] = 'Class:'
    ws['B4'] = class_obj.name
    ws['A4'].font = Font(bold=True, size=10)

    semester_label = f"Semester {semester} ({start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')})"
    ws['A5'] = 'Period:'
    ws['B5'] = semester_label
    ws['A5'].font = Font(bold=True, size=10)

    ws['A6'] = 'Total Students:'
    ws['B6'] = len(summaries)
    ws['A6'].font = Font(bold=True, size=10)

    # Table headers
    table_header_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
    table_header_font = Font(bold=True, color='FFFFFF', size=10)
    table_header_alignment = Alignment(horizontal='center', vertical='center')

    headers = ['No.', 'Student Name', 'Total Days', 'Present', 'Absent', 'Sick', 'Attendance %']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = table_header_alignment

    # Data rows with color coding
    border_thin = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb'),
    )

    for row_idx, summary in enumerate(summaries, 9):
        ws.cell(row=row_idx, column=1, value=row_idx - 8)
        ws.cell(row=row_idx, column=2, value=summary['student'].get_full_name())
        ws.cell(row=row_idx, column=3, value=summary['total_days'])
        ws.cell(row=row_idx, column=4, value=summary['present'])
        ws.cell(row=row_idx, column=5, value=summary['absent'])
        ws.cell(row=row_idx, column=6, value=summary['sick'])
        ws.cell(row=row_idx, column=7, value=f"{summary['percentage']:.1f}%")

        # Color by attendance percentage
        attendance_pct = summary['percentage']
        if attendance_pct >= 90:
            row_color = 'dcfce7'  # Green
        elif attendance_pct >= 80:
            row_color = 'fef3c7'  # Amber
        else:
            row_color = 'fee2e2'  # Red

        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center')
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left')

    workbook.save(response)
    return response
