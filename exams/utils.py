"""Utility functions for exam operations"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def export_exam_results_to_xlsx(exam):
    """
    Export exam results to Excel workbook
    Returns BytesIO object with xlsx file
    Format: Nama Siswa | NIS | Skor | Waktu Selesai
    """
    from academic.models import StudentEnrollment

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Hasil Ujian'

    # Header styling
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    headers = ['Nama Siswa', 'NIS', 'Skor', 'Waktu Selesai']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Get exam sessions that are completed
    sessions = exam.sessions.filter(status__in=['SUBMITTED', 'GRADED']).select_related('student').order_by('-submitted_at')

    # Write data rows
    for row_num, session in enumerate(sessions, 2):
        # Student Name
        cell = sheet.cell(row=row_num, column=1)
        cell.value = session.student.get_full_name()
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

        # NIS - ambil dari StudentEnrollment.student_id_in_class
        cell = sheet.cell(row=row_num, column=2)
        nis = '-'

        # Get enrollment for this student in the exam's class
        enrollment = StudentEnrollment.objects.filter(
            student=session.student,
            class_obj=exam.class_obj
        ).first()

        if enrollment and enrollment.student_id_in_class:
            nis = enrollment.student_id_in_class
        else:
            # Fallback ke UserProfile.nis jika ada
            if hasattr(session.student, 'profile') and session.student.profile.nis:
                nis = session.student.profile.nis
            else:
                # Fallback terakhir ke username
                nis = session.student.username

        cell.value = nis
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

        # Score
        cell = sheet.cell(row=row_num, column=3)
        cell.value = session.score or 0
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Submitted At (Waktu Selesai)
        cell = sheet.cell(row=row_num, column=4)
        cell.value = session.submitted_at.strftime('%d/%m/%Y %H:%M') if session.submitted_at else '-'
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')

    # Adjust column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 18

    # Freeze header row
    sheet.freeze_panes = 'A2'

    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output
