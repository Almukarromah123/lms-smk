from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.http import JsonResponse, Http404, HttpResponse
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.contrib import messages
from django.db.models import Max
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import json
import random
import uuid
from .forms import ExamQuestionForm, ExamQuestionImportForm
from .models import Exam, ExamSession, ExamQuestion
from .utils import export_exam_results_to_xlsx


class StudentExamListView(LoginRequiredMixin, ListView):
    """List exams for student - only shows exams on their scheduled date that student hasn't completed"""
    template_name = 'exams/student_list.html'
    context_object_name = 'exams'
    paginate_by = 20

    def get_queryset(self):
        from django.utils import timezone
        from django.db.models import Exists, OuterRef
        from datetime import timedelta
        
        # Get timezone aware start and end of today in UTC
        # Using replace() to keep UTC timezone and avoid timezone conversion issues
        now = timezone.now()
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        student = self.request.user
        
        # Only show exams for today's date using UTC datetime range
        # This works correctly with timezone aware datetime in MySQL
        exams = Exam.objects.filter(
            is_published=True,
            exam_date__gte=start_of_day,
            exam_date__lt=end_of_day
        )
        
        # Exclude exams where student already submitted/completed
        completed_exams = ExamSession.objects.filter(
            student=student,
            exam=OuterRef('id'),
            status__in=['SUBMITTED', 'GRADED']
        )
        exams = exams.exclude(Exists(completed_exams))
        
        return exams.order_by('exam_date')


class ExamDetailView(LoginRequiredMixin, DetailView):
    """View exam details"""
    model = Exam
    template_name = 'exams/detail.html'
    context_object_name = 'exam'
    pk_url_kwarg = 'exam_id'
    
    def get_context_data(self, **kwargs):
        from django.utils import timezone
        from academic.models import StudentEnrollment
        from django.contrib import messages
        
        context = super().get_context_data(**kwargs)
        exam = self.get_object()
        today = timezone.now().date()
        
        is_enrolled = StudentEnrollment.objects.filter(
            student=self.request.user,
            class_obj=exam.class_obj,
            status='ACTIVE'
        ).exists()
        context['is_enrolled'] = is_enrolled
        
        # Check if exam is available for students today
        if self.request.user.role == 'STUDENT':
            # Check if today is exam date
            is_exam_today = exam.exam_date.date() == today
            context['is_exam_today'] = is_exam_today
            
            # Check if student already completed this exam
            already_submitted = ExamSession.objects.filter(
                student=self.request.user,
                exam=exam,
                status__in=['SUBMITTED', 'GRADED']
            ).exists()
            context['already_submitted'] = already_submitted
            
            # Check if exam date has passed
            exam_date_passed = exam.exam_date.date() < today
            context['exam_date_passed'] = exam_date_passed
        
        return context


class StartExamView(LoginRequiredMixin, DetailView):
    """Start exam - creates/gets exam session and redirects to session"""
    model = Exam
    template_name = None  # No template, redirects instead
    pk_url_kwarg = 'exam_id'

    def get(self, request, *args, **kwargs):
        """Create or get exam session and redirect"""
        from django.utils import timezone
        from django.contrib import messages
        from academic.models import StudentEnrollment
        
        exam = self.get_object()
        student = request.user
        today = timezone.now().date()

        # Check if student is enrolled in the class for this exam
        enrollment = StudentEnrollment.objects.filter(
            student=student,
            class_obj=exam.class_obj,
            status='ACTIVE'
        ).first()

        if not enrollment:
            messages.error(request, "Anda tidak terdaftar di kelas ini.")
            return redirect('exams:student_list')

        # Check if exam date is today
        if exam.exam_date.date() != today:
            messages.error(request, "Ujian ini hanya tersedia pada tanggal yang dijadwalkan.")
            return redirect('exams:student_list')

        # Check if student already submitted this exam
        existing_submission = ExamSession.objects.filter(
            student=student,
            exam=exam,
            status__in=['SUBMITTED', 'GRADED']
        ).first()
        
        if existing_submission:
            messages.error(request, "Anda sudah menyelesaikan ujian ini.")
            return redirect('exams:student_list')

        # Create or get exam session
        session, created = ExamSession.objects.get_or_create(
            exam=exam,
            student=student
        )

        # Redirect to exam session
        return redirect('exams:session', session_id=session.id)

    def get_queryset(self):
        """Get all exams - published status controlled by is_published field in edit"""
        return Exam.objects.all()


class ExamSessionView(LoginRequiredMixin, DetailView):
    """Take exam session"""
    model = ExamSession
    template_name = 'exams/session.html'
    context_object_name = 'session'
    pk_url_kwarg = 'session_id'

    def get_object(self, queryset=None):
        """Get exam session ensuring student owns it"""
        session_id = self.kwargs.get('session_id')

        try:
            session = ExamSession.objects.get(id=session_id, student=self.request.user)
        except ExamSession.DoesNotExist:
            raise Http404("Exam session not found or access denied")

        if session.status == 'NOT_STARTED':
            session.start_exam()

        if session.status == 'IN_PROGRESS' and session.is_time_up():
            session.submit_exam()
            messages.info(self.request, "Waktu ujian habis, jawaban otomatis terkumpul.")

        return session

    def post(self, request, *args, **kwargs):
        session = self.get_object()

        if session.status != 'IN_PROGRESS':
            messages.warning(request, "Ujian sudah selesai.")
            return redirect('exams:session', session_id=session.id)

        if session.is_time_up():
            session.submit_exam()
            messages.info(request, "Waktu ujian habis, jawaban otomatis terkumpul.")
            return redirect('exams:session', session_id=session.id)

        answers = {}
        for question in session.exam.questions.filter(is_active=True).order_by('order'):
            answer = request.POST.get(f'question_{question.id}', '').strip()
            if answer != '':
                answers[str(question.id)] = answer

        session.student_answers = answers
        session.save()

        if request.POST.get('submit_exam'):
            session.submit_exam()
            session.save()
            messages.success(request, 'Ujian selesai. Jawaban telah terkumpul.')
            return redirect('exams:session', session_id=session.id)

        messages.success(request, 'Jawaban telah disimpan sementara.')
        return redirect('exams:session', session_id=session.id)

    def get_context_data(self, **kwargs):
        """Add exam questions to context"""
        context = super().get_context_data(**kwargs)
        session = self.get_object()

        questions = session.exam.questions.filter(is_active=True).order_by('order')
        if session.exam.shuffle_questions:
            questions = list(questions)
            random.shuffle(questions)

        context['questions'] = questions
        context['saved_answers'] = session.student_answers or {}
        context['time_remaining_seconds'] = int(session.get_time_remaining_seconds())
        context['question_per_page'] = session.exam.question_per_page
        context['allow_back'] = session.exam.allow_back
        context['question_count'] = len(questions)
        return context


class TeacherExamListView(LoginRequiredMixin, ListView):
    """List exams created by teacher"""
    template_name = 'exams/teacher_list.html'
    context_object_name = 'exams'
    paginate_by = 20

    def get_queryset(self):
        return Exam.objects.filter(created_by=self.request.user).order_by('-exam_date')


class CreateExamView(LoginRequiredMixin, CreateView):
    """Create new exam"""
    model = Exam
    template_name = 'exams/create.html'
    fields = ['class_obj', 'subject', 'title', 'description', 'exam_date', 'duration_minutes', 'total_points', 'shuffle_questions', 'shuffle_options', 'allow_back', 'question_per_page', 'display_answer_key', 'show_feedback']

    def get_initial(self):
        """Set default exam_date to today at 17:00"""
        from datetime import datetime, timezone
        initial = super().get_initial()
        today = datetime.now(timezone.utc).replace(hour=17, minute=0, second=0, microsecond=0)
        initial['exam_date'] = today
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.is_published = True
        messages.success(self.request, 'Ujian berhasil dibuat! Sekarang tambahkan soal-soal ujian.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('exams:questions', kwargs={'exam_id': self.object.id})


class EditExamView(LoginRequiredMixin, UpdateView):
    """Edit exam"""
    model = Exam
    template_name = 'exams/edit.html'
    fields = ['title', 'description', 'exam_date', 'duration_minutes', 'total_points', 'shuffle_questions', 'shuffle_options', 'allow_back', 'question_per_page', 'display_answer_key', 'show_feedback', 'is_published']
    pk_url_kwarg = 'exam_id'
    
    def form_valid(self, form):
        messages.success(self.request, 'Ujian berhasil diperbarui!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('exams:questions', kwargs={'exam_id': self.object.id})


class ManageQuestionsView(LoginRequiredMixin, DetailView):
    """Manage exam questions"""
    model = Exam
    template_name = 'exams/questions.html'
    context_object_name = 'exam'
    pk_url_kwarg = 'exam_id'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        exam = self.get_object()
        context['questions'] = exam.questions.filter(is_active=True).order_by('order')
        context['import_form'] = None
        return context


class AddQuestionView(LoginRequiredMixin, CreateView):
    model = ExamQuestion
    form_class = ExamQuestionForm
    template_name = 'exams/question_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.exam = get_object_or_404(Exam, id=self.kwargs['exam_id'], created_by=request.user)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.exam = self.exam
        last_order = self.exam.questions.aggregate(max_order=Max('order'))['max_order']
        form.instance.order = (last_order + 1) if last_order is not None else 0

        # Handle image uploads for options A-D
        options_images = {}
        for option_code in ['A', 'B', 'C', 'D']:
            file_key = f'option_image_{option_code}'
            if file_key in self.request.FILES:
                image_file = self.request.FILES[file_key]
                if image_file:
                    # Save file to media directory with unique name
                    file_path = default_storage.save(
                        f'exam_options/{uuid.uuid4()}_{image_file.name}',
                        ContentFile(image_file.read())
                    )
                    options_images[option_code] = file_path

        if options_images:
            form.instance.options_images = options_images

        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('exams:questions', kwargs={'exam_id': self.exam.id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['exam'] = self.exam
        context['title'] = 'Add Question'
        return context


class EditQuestionView(LoginRequiredMixin, UpdateView):
    model = ExamQuestion
    form_class = ExamQuestionForm
    template_name = 'exams/question_form.html'
    pk_url_kwarg = 'question_id'

    def get_queryset(self):
        return ExamQuestion.objects.filter(exam__created_by=self.request.user)

    def form_valid(self, form):
        # Handle image uploads for options A-D
        options_images = form.instance.options_images or {}

        for option_code in ['A', 'B', 'C', 'D']:
            file_key = f'option_image_{option_code}'
            if file_key in self.request.FILES:
                image_file = self.request.FILES[file_key]
                if image_file:
                    # Save file to media directory with unique name
                    file_path = default_storage.save(
                        f'exam_options/{uuid.uuid4()}_{image_file.name}',
                        ContentFile(image_file.read())
                    )
                    options_images[option_code] = file_path

        if options_images:
            form.instance.options_images = options_images

        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('exams:questions', kwargs={'exam_id': self.object.exam.id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['exam'] = self.object.exam
        context['title'] = 'Edit Question'
        return context


class DeleteQuestionView(LoginRequiredMixin, DeleteView):
    model = ExamQuestion
    template_name = 'exams/question_confirm_delete.html'
    pk_url_kwarg = 'question_id'

    def get_queryset(self):
        return ExamQuestion.objects.filter(exam__created_by=self.request.user)

    def get_success_url(self):
        return reverse_lazy('exams:questions', kwargs={'exam_id': self.object.exam.id})


class ImportQuestionsView(LoginRequiredMixin, FormView):
    template_name = 'exams/import_questions.html'
    form_class = ExamQuestionImportForm

    def dispatch(self, request, *args, **kwargs):
        self.exam = get_object_or_404(Exam, id=self.kwargs['exam_id'], created_by=request.user)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['exam'] = self.exam
        return context

    def form_valid(self, form):
        workbook = load_workbook(filename=form.cleaned_data['xlsx_file'], data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows or len(rows) < 2:
            form.add_error('xlsx_file', 'File tidak berisi data yang valid atau hanya ada header.')
            return self.form_invalid(form)

        headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
        required_fields = ['question_type', 'question_text', 'points', 'options_data', 'correct_answer']
        missing = [field for field in required_fields if field not in headers]
        if missing:
            form.add_error('xlsx_file', f'Header wajib hilang: {", ".join(missing)}')
            return self.form_invalid(form)

        created = 0
        current_order = self.exam.questions.aggregate(max_order=Max('order'))['max_order']
        if current_order is None:
            current_order = -1

        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue

            values = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            question_text = values.get('question_text')
            if not question_text:
                continue

            options_data_raw = values.get('options_data') or '{}'
            if isinstance(options_data_raw, str) and options_data_raw.strip():
                try:
                    options_data = json.loads(options_data_raw)
                except json.JSONDecodeError:
                    options_data = {}
            else:
                options_data = {}

            current_order += 1
            ExamQuestion.objects.create(
                exam=self.exam,
                question_type=str(values.get('question_type') or 'MCQ').upper(),
                question_text=str(question_text),
                points=float(values.get('points') or 1),
                options_data=options_data,
                correct_answer=str(values.get('correct_answer') or '').strip(),
                explanation=str(values.get('explanation') or '').strip(),
                order=current_order
            )
            created += 1

        messages.success(self.request, f'{created} soal berhasil diimpor ke ujian.')
        return redirect('exams:questions', exam_id=self.exam.id)


class ExamResultsView(LoginRequiredMixin, DetailView):
    """View exam results & export"""
    model = Exam
    template_name = 'exams/results.html'
    context_object_name = 'exam'
    pk_url_kwarg = 'exam_id'

    def get_queryset(self):
        """Only allow teacher who created exam or admin to view/export results"""
        user = self.request.user
        if user.role == 'ADMIN':
            return Exam.objects.all()
        return Exam.objects.filter(created_by=user)

    def get(self, request, *args, **kwargs):
        """Handle export request"""
        exam = self.get_object()

        # Check permission
        if request.user.role != 'ADMIN' and exam.created_by != request.user:
            return HttpResponse('Unauthorized', status=403)

        # Check if export requested
        export_format = request.GET.get('format')
        if export_format == 'xlsx':
            # Generate Excel file
            file_obj = export_exam_results_to_xlsx(exam)

            # Create response
            response = HttpResponse(
                file_obj.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"exam_results_{exam.id}_{exam.title.replace(' ', '_')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # Normal view
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        exam = self.get_object()
        context['sessions'] = exam.sessions.select_related('student').order_by('-started_at')
        context['can_export'] = exam.created_by == self.request.user or self.request.user.role == 'ADMIN'
        return context


class DeleteExamView(LoginRequiredMixin, DeleteView):
    """Delete exam"""
    model = Exam
    template_name = 'exams/delete_confirm.html'
    pk_url_kwarg = 'exam_id'
    success_url = reverse_lazy('exams:teacher_list')

    def get_queryset(self):
        """Only allow teacher to delete own exams"""
        return Exam.objects.filter(created_by=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Exam berhasil dihapus!')
        return super().delete(request, *args, **kwargs)


@require_http_methods(["POST"])
def log_violation_api(request):
    """API endpoint to log exam proctoring violations"""
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        violation_type = data.get('violation_type')
        tab_switches = data.get('tab_switches', 0)

        # Get session
        session = ExamSession.objects.get(id=session_id, student=request.user)

        # Initialize or update violations log
        if not session.violations_log:
            session.violations_log = []

        violation_record = {
            'type': violation_type,
            'timestamp': json.dumps({
                'iso': str(json.loads(json.dumps(str(timezone.now())))),
                'timestamp': str(timezone.now())
            }).replace('\\', ''),
            'tab_switches': tab_switches
        }

        session.violations_log.append(violation_record)
        session.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Violation logged',
            'violation_count': len(session.violations_log)
        })

    except ExamSession.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Session not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
